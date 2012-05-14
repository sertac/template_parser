'''
Created on 13 May 2012

@author: sertac
'''

from django.core.management.base import BaseCommand, CommandError
from openpyxl.reader.excel import load_workbook


class Command(BaseCommand):
    args = '<file file ...>'
    help = 'parses templated excel files'
    cells_dict = {}

    def handle(self, *args, **options):
        for file in args:
            wb = load_workbook(filename=args[0], use_iterators=True)
            sheets = wb.get_sheet_names()
            print sheets
            ws = wb.get_sheet_by_name(sheets[2])
            rows = []
            headers = []
            data = []
            extra = []

            for row in ws.iter_rows():
                rows.append(row)

            for row in rows:
                for rawcell in row:
                    if rawcell.internal_value == "Country":
                        headers.append(row)

            for row in rows:
                if row[0].row < headers[0][0].row:
                    extra.append(row)
                elif row[0].row > headers[0][0].row:
                    data.append(row)

        field_names = []
        for row in headers:
            for raw_cell in row:
                field_names.append(raw_cell.internal_value)

        field_extra = {}
        for row in extra:
            row_data = []
            for raw_cell in row:
                if raw_cell is not None:
                    row_data.append(raw_cell.internal_value)
                    field_extra[raw_cell.row] = row_data

        field_values = {}
        for row in data:
            row_data = []
            for raw_cell in row:
                if raw_cell is not None:
                    row_data.append(raw_cell.internal_value)
                    field_values[raw_cell.row] = row_data

        print field_extra[2]
        print field_names
        print field_values[6]
